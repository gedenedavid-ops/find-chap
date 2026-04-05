import pandas as pd
import openpyxl

file_path = r"C:\Users\geden\Desktop\find-chap\RESULTATS DEFINITIFS_SEMESTRE 4_SESSION 2.xlsx"

print("=== PANDAS READ (default header=0) ===")
df = pd.read_excel(file_path)
print(f"Shape: {df.shape}")
print(f"Columns: {df.columns.tolist()}")
print("\nFirst 10 rows:")
print(df.head(10))

print("\n" + "="*80)
print("\n=== OPENPYXL (RAW CELLS) ===")
wb = openpyxl.load_workbook(file_path)
ws = wb.active
print(f"Sheet name: {ws.title}")
print(f"Max row: {ws.max_row}, Max col: {ws.max_column}")

print("\nFirst 15 rows (raw):")
for i in range(1, min(16, ws.max_row + 1)):
    row_values = []
    for j in range(1, min(10, ws.max_column + 1)):
        cell = ws.cell(i, j)
        row_values.append(cell.value)
    print(f"Row {i}: {row_values}")
